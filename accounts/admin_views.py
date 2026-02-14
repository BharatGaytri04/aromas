from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.utils import timezone
from .models import Account, LoginAttempt
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@staff_member_required
def view_all_users(request):
    """View all users in admin panel"""
    users = Account.objects.all().order_by('-date_joined')
    
    # Get login statistics for each user
    user_data = []
    for user in users:
        login_attempts = LoginAttempt.objects.filter(user=user).count()
        successful_logins = LoginAttempt.objects.filter(user=user, success=True).count()
        last_login_attempt = LoginAttempt.objects.filter(user=user).order_by('-attempted_at').first()
        
        user_data.append({
            'user': user,
            'login_attempts': login_attempts,
            'successful_logins': successful_logins,
            'last_login_attempt': last_login_attempt.attempted_at if last_login_attempt else None,
        })
    
    context = {
        'users': user_data,
        'total_users': users.count(),
        'active_users': users.filter(is_active=True).count(),
        'inactive_users': users.filter(is_active=False).count(),
    }
    
    return render(request, 'admin/view_users.html', context)


@staff_member_required
def export_users_excel(request):
    """Export all users to Excel file"""
    users = Account.objects.all().order_by('-date_joined')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Users"
    
    # Header row with styling
    headers = [
        'ID', 'Email', 'Username', 'First Name', 'Last Name', 'Phone Number',
        'Address Line 1', 'Address Line 2', 'City', 'State', 'Country', 'Pincode',
        'Date Joined', 'Last Login', 'Is Active', 'Is Staff', 'Is Admin',
        'Total Login Attempts', 'Successful Logins'
    ]
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add user data
    for row_num, user in enumerate(users, 2):
        login_attempts = LoginAttempt.objects.filter(user=user).count()
        successful_logins = LoginAttempt.objects.filter(user=user, success=True).count()
        
        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=user.email)
        ws.cell(row=row_num, column=3, value=user.username)
        ws.cell(row=row_num, column=4, value=user.first_name)
        ws.cell(row=row_num, column=5, value=user.last_name)
        ws.cell(row=row_num, column=6, value=user.phone_number)
        ws.cell(row=row_num, column=7, value=user.address_line_1)
        ws.cell(row=row_num, column=8, value=user.address_line_2)
        ws.cell(row=row_num, column=9, value=user.city)
        ws.cell(row=row_num, column=10, value=user.state)
        ws.cell(row=row_num, column=11, value=user.country)
        ws.cell(row=row_num, column=12, value=user.pincode)
        ws.cell(row=row_num, column=13, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '')
        ws.cell(row=row_num, column=14, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')
        ws.cell(row=row_num, column=15, value='Yes' if user.is_active else 'No')
        ws.cell(row=row_num, column=16, value='Yes' if user.is_staff else 'No')
        ws.cell(row=row_num, column=17, value='Yes' if user.is_admin else 'No')
        ws.cell(row=row_num, column=18, value=login_attempts)
        ws.cell(row=row_num, column=19, value=successful_logins)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))
        for row_num in range(2, len(users) + 2):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="all_users_{timestamp}.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
def export_users_csv(request):
    """Export all users to CSV file (fallback if Excel not available)"""
    users = Account.objects.all().order_by('-date_joined')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="all_users_{timestamp}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Email', 'Username', 'First Name', 'Last Name', 'Phone Number',
        'Address Line 1', 'Address Line 2', 'City', 'State', 'Country', 'Pincode',
        'Date Joined', 'Last Login', 'Is Active', 'Is Staff', 'Is Admin',
        'Total Login Attempts', 'Successful Logins'
    ])
    
    for user in users:
        login_attempts = LoginAttempt.objects.filter(user=user).count()
        successful_logins = LoginAttempt.objects.filter(user=user, success=True).count()
        
        writer.writerow([
            user.id,
            user.email,
            user.username,
            user.first_name,
            user.last_name,
            user.phone_number,
            user.address_line_1,
            user.address_line_2,
            user.city,
            user.state,
            user.country,
            user.pincode,
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
            'Yes' if user.is_active else 'No',
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_admin else 'No',
            login_attempts,
            successful_logins,
        ])
    
    return response
